from openpyxl import Workbook # type: ignore
from openpyxl.styles import Font # type: ignore
from openpyxl.styles import Alignment #type:ignore
from openpyxl.styles import PatternFill#type:ignore

workbook = Workbook()

sheet = workbook.active
sheet.title = "Employees"
sheet = workbook["Employees"]
#Heading

#mergecell
sheet.merge_cells("A1:D1")

sheet["A1"] = "ABC Company Pvt Ltd \
\n Employee Salary Report"

sheet["A1"].font = Font(
    bold=True,
    italic=True,
    size=18,
    color="FF0000"
)
# set title and header alignments
sheet["A1"].alignment = Alignment(horizontal="center", vertical="center")

# Header  bg Color 
fill = PatternFill(
    fill_type="solid",
    fgColor="FFFF00"
)

sheet["A1"].fill = fill

# Sub heading row A2:D2 font style 
sheet.append(["ID", "Name", "Department", "Salary"])
for cell in sheet["A2:D2"][0]:
    cell.font = Font(bold="true", size=12)

# center align the header row A2:D2
for cell in sheet["A2:D2"][0]:
    cell.alignment = Alignment(horizontal="center", vertical="center")

# Data
sheet.append([101, "Boopathi", "IT", 50000])
sheet.append([102, "Ram", "HR", 45000])
sheet.append([103,"Arjun","Admin",30000])

for row in range(3, 6):      # D3 to D5 (salary rows)
    sheet[f"D{row}"].number_format = '#,##0'


#Freeze sub heading 
sheet.freeze_panes = "A2"
workbook.save("salary_report.xlsx")

print("Excel Created Successfully")