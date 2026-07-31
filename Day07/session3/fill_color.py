from openpyxl import Workbook #type:ignore
from openpyxl.styles import PatternFill #type:ignore

wb = Workbook()
sheet = wb.active

sheet["A1"] = "Employee Name"

fill = PatternFill(
    fill_type="solid",
    fgColor="FFFF00"
)

sheet["A1"].fill = fill

wb.save("fill_demo.xlsx")