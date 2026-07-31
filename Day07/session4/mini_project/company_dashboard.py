from openpyxl import Workbook # type: ignore
from openpyxl.styles import Font # type: ignore
from openpyxl.styles import Alignment #type:ignore
from openpyxl.styles import PatternFill#type:ignore
from openpyxl.chart import BarChart, PieChart, Reference #type:ignore

workbook = Workbook()

sheet = workbook.active
sheet.title = "Employees"
sheet = workbook["Employees"]
#Heading

#mergecell
sheet.merge_cells("A1:D1")

sheet["A1"] = "ABC Company Pvt Ltd \
\n Employee Salary Report"

sheet["A1"].font = Font(
    bold=True,
    italic=True,
    size=18,
    color="FF0000"
)
# set title and header alignments
sheet["A1"].alignment = Alignment(horizontal="center", vertical="center")

# Header  bg Color 
fill = PatternFill(
    fill_type="solid",
    fgColor="FFFF00"
)

sheet["A1"].fill = fill

# Sub heading row A2:D2 font style 
sheet.append(["ID", "Name", "Department", "Salary"])
for cell in sheet["A2:D2"][0]:
    cell.font = Font(bold="true", size=12)

# center align the header row A2:D2
for cell in sheet["A2:D2"][0]:
    cell.alignment = Alignment(horizontal="center", vertical="center")

# Data
sheet.append(["Department", "Salary"])

sheet.append(["IT",80000])
sheet.append(["HR",45000])
sheet.append(["Sales",55000])
sheet.append(["Admin",35000])


for row in range(3, 6):      # D3 to D5 (salary rows)
    sheet[f"D{row}"].number_format = '#,##0'


#Freeze sub heading 
sheet.freeze_panes = "A2"

data = Reference(sheet, min_col=2, min_row=1, max_row=5)
labels = Reference(sheet, min_col=1, min_row=2, max_row=5)

bar = BarChart()
bar.title = "Salary"

bar.add_data(data, titles_from_data=True)
bar.set_categories(labels)

pie = PieChart()
pie.title = "Distribution"

pie.add_data(data, titles_from_data=True)
pie.set_categories(labels)

sheet.add_chart(bar, "D2")
sheet.add_chart(pie, "D18")

workbook.save("company_dashboard.xlsx")

print("Excel Created Successfully")