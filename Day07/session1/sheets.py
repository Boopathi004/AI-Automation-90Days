from openpyxl import Workbook # type: ignore

workbook = Workbook()

sheet1 = workbook.active
sheet1.title = "Employees"

workbook.create_sheet("Sales")
workbook.create_sheet("HR")
workbook.create_sheet("Finance")

workbook.save("company.xlsx")

print("Workbook with multiple sheets created!")