from openpyxl import load_workbook # type: ignore

workbook = load_workbook("company.xlsx")

print(workbook.sheetnames)