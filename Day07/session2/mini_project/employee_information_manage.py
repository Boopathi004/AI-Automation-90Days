from openpyxl import load_workbook #type: ignore
import sys

workbook=load_workbook("Employee_mini.xlsx")

sheet=workbook["Employees"]

print("=" *40)
print("Employees information management system " )
print("=" *40)

print("\n""select your option" \
"\n1.View All Employee Details ," \
"\n2.Search Employee, " \
"\n3.Update Salary Details" \
"\n press any key to exit... "
"\n")
choice=int(input("Enter Your choice "))
#view All the Employee Details

if (choice == 1 ):
    for row in sheet.iter_rows(values_only=True):
     print(row)

             
   # search Employee 
elif (choice == 2 ):
   employee_id = int(input("Enter Employee ID: "))
   found = False
   for row in sheet.iter_rows(min_row=2):
    if row[0].value == employee_id:
        print("Employee Found")
        print("Name :", row[1].value)
        print("Department :", row[2].value)
        print("Salary :", row[3].value)
        found = True
        break
    if not found:
     print("Employee Not Found")

   #update salary      
elif(choice == 3):
  employee_id = int(input("Enter Employee ID: "))
  salary=int(input("Enter New Salary : "))
  found1=False
  for row in sheet.iter_rows(min_row=2):
    if row[0].value == employee_id:
        row[3].value=salary
        #sheet["D2"] = 65000
        print("increment_Salary :", row[3].value)
        workbook.save("Employee_mini.xlsx")
        print("Salary Updated Successfully")
        found1=True
    if not found1:
        print("Employee Not Found")
else:
  sys.exit()



 






