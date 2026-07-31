from openpyxl import load_workbook #type:ignore

workbook = load_workbook("Employee_mini.xlsx")

sheet = workbook["Employees"]

for col in sheet.iter_cols(values_only=True):
    print(col)