from openpyxl import load_workbook #type:ignore

workbook = load_workbook("Employee_mini.xlsx")

sheet = workbook["Employees"]

for row in sheet.iter_rows(values_only=True):
    print(row)