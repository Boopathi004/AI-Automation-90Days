from openpyxl import load_workbook #type:ignore

workbook = load_workbook("Employee_mini.xlsx")

sheet = workbook["Employees"]

employee_id = 102

found = False

for row in sheet.iter_rows(min_row=2):
    if row[0].value == employee_id:
        print("Employee Found")
        print("Name :", row[1].value)
        print("Department :", row[2].value)
        print("Salary :", row[3].value)
        found = True
        break

if not found:
    print("Employee Not Found")